#!/usr/bin/env python3
"""
update-work-log.py — append a new work-session row to the master work-log Excel.

Source of truth: docs/observability/QA_Nexus_Work_Log.xlsx (Kimi-redesigned "Professional"
layout swapped in 2026-05-06 Day 11; backfilled with all 26 missing rows from the prior
layout — see chore/work-log-professional-swap PR).

Layout note (Day-11 swap):
  - Rows 1-4 are banner/title/subtitle/blank
  - Row 5 is the HEADER row
  - Row 6+ is data
  - Last row is GRAND TOTAL (with =SUM() formulas the script re-extends on insert)
  - Column A is an empty gutter; data starts at column B
  - "All Sessions" sheet has 9 columns of data (B-J) including a "Parallel Work (FE+BE)"
    column populated by the parallel-work correlation hook (Day-11 TASK 6, deferred to Thu)

This script:
  1. Validates input args + loads the workbook
  2. Appends a row to "All Sessions" sheet (and the per-phase sheet)
  3. Re-extends GRAND TOTAL =SUM() formulas to cover the new row
  4. Saves the workbook in-place (no separate output file — same spreadsheet stays canonical)

Usage:
    pnpm work:log -- \\
        --date 2026-05-01 \\
        --day Friday \\
        --start "10:00 AM" \\
        --end "11:30 AM" \\
        --hours 1.50 \\
        --files 12 \\
        --phase Development \\
        --theme "OTel SDK wiring + smoke test" \\
        --what "Wired OTel SDK in apps/api/src/observability/otel.config.ts; first trace visible in Grafana Cloud."

Or dry-run (prints what would be appended without saving):
    pnpm work:log -- --date 2026-05-01 --day Friday --start "10:00 AM" --end "11:30 AM" --hours 1.5 --files 12 --phase Development --theme "..." --what "..." --dry-run

Phases (must match a sheet name in the workbook):
    "Idea Confirmation" | "Test Case Management Research" | "Test Automation & Reporting"
    | "Design & Specifications" | "Milestone Planning" | "Launch Creative" | "Development" | "Other"

Cross-references:
  - docs/eod-reports/2026-04-30-day-4.md — the consolidation that established this aggregator's contract
  - docs/observability/QA_Nexus_Work_Log.xlsx — the workbook this script writes to
"""

from __future__ import annotations

import argparse
import sys
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:
    sys.stderr.write(
        "ERROR: openpyxl not installed. Run: pip install openpyxl --break-system-packages\n"
    )
    sys.exit(1)

# Paths — resolved from this script's location for portability across worktrees
THIS = Path(__file__).resolve()
REPO_ROOT = THIS.parent.parent
WORKBOOK_PATH = REPO_ROOT / "docs" / "observability" / "QA_Nexus_Work_Log.xlsx"

VALID_PHASES = {
    "Idea Confirmation",
    "Test Case Management Research",
    "Test Automation & Reporting",
    "Design & Specifications",
    "Milestone Planning",
    "Launch Creative",
    "Development",
    "Other",
}

# Style constants (must match what's already in the workbook for visual consistency)
THIN = Side(style="thin", color="D1D5DB")
BR = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
CTR = Alignment(horizontal="center", vertical="center", wrap_text=True)
NORM_FONT = Font(name="Calibri", size=10)
DAY_TOTAL_FONT = Font(name="Calibri", size=10, bold=True, color="0F766E")
DAY_TOTAL_FILL = PatternFill("solid", start_color="F0FDFA")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Append a work-session row to QA_Nexus_Work_Log.xlsx")
    p.add_argument("--date", required=True, help="Session date YYYY-MM-DD (e.g. 2026-05-01)")
    p.add_argument("--day", required=True, help="Day-of-week label (Monday/Tuesday/...)")
    p.add_argument("--start", required=True, help="Start time, free form (e.g. '10:00 AM')")
    p.add_argument("--end", required=True, help="End time, free form (e.g. '11:30 AM')")
    p.add_argument(
        "--hours",
        required=True,
        type=float,
        help="Duration in decimal hours (e.g. 1.50 for 1h30m)",
    )
    p.add_argument(
        "--files",
        required=True,
        type=int,
        help="Files-touched count (any granularity — file-paths or commit shas, your call)",
    )
    p.add_argument(
        "--phase",
        required=True,
        choices=sorted(VALID_PHASES),
        help="Phase / theme bucket (must match an existing sheet name)",
    )
    p.add_argument(
        "--theme",
        required=True,
        help="Short phrase for the sample/theme column (e.g. 'OTel SDK wiring')",
    )
    p.add_argument(
        "--what",
        required=True,
        help="What I did — narrative paragraph going into the 'What I Did' column",
    )
    p.add_argument(
        "--dry-run",
        action="store_true",
        help="Print what would be appended without saving the workbook",
    )
    return p.parse_args()


def validate(args: argparse.Namespace) -> None:
    """Validate input args. Exit non-zero on validation error."""
    try:
        datetime.strptime(args.date, "%Y-%m-%d")
    except ValueError:
        sys.stderr.write(f"ERROR: --date must be YYYY-MM-DD format, got '{args.date}'\n")
        sys.exit(2)
    if args.hours <= 0 or args.hours > 24:
        sys.stderr.write(f"ERROR: --hours must be in (0, 24], got {args.hours}\n")
        sys.exit(2)
    if args.files < 0:
        sys.stderr.write(f"ERROR: --files cannot be negative, got {args.files}\n")
        sys.exit(2)
    if not WORKBOOK_PATH.exists():
        sys.stderr.write(
            f"ERROR: workbook not found at {WORKBOOK_PATH}\n"
            f"  Expected: docs/observability/QA_Nexus_Work_Log.xlsx (per Day-4 consolidation)\n"
        )
        sys.exit(3)


# Day-11 swap: NEW layout has empty col A (gutter) — data starts at col B.
# To convert from OLD (col 1=Date) to NEW (col 2=Date), we add COL_OFFSET to every column index.
COL_OFFSET = 1


def _find_grand_total_row(ws):
    """Find the row containing 'GRAND TOTAL' in any column. Returns row index or None."""
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if isinstance(v, str) and v.strip().upper() == "GRAND TOTAL":
                return r
    return None


def _extend_sum_formulas(ws, header_row: int, gt_row_after_insert: int) -> int:
    """Re-extend any =SUM(X#:Y#) formula on the GRAND TOTAL row to cover [header_row+1, gt_row-1].

    Returns the count of formulas updated. Day-11 swap added this; OLD layout was Excel-recompute-only.
    """
    import re
    pat = re.compile(r"=SUM\(([A-Z]+)\d+:([A-Z]+)\d+\)", re.IGNORECASE)
    new_first = header_row + 1
    new_last = gt_row_after_insert - 1
    updates = 0
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=gt_row_after_insert, column=c).value
        if isinstance(v, str) and v.upper().startswith("=SUM("):
            m = pat.match(v)
            if m:
                a, b = m.group(1), m.group(2)
                new = f"=SUM({a}{new_first}:{b}{new_last})"
                if new != v:
                    ws.cell(row=gt_row_after_insert, column=c).value = new
                    updates += 1
    return updates


def append_to_phase_sheet(wb, args: argparse.Namespace) -> int:
    """Append the row to the phase-specific sheet. Returns the new row index.

    Day-11 layout: header at row 5; data starts row 6; GRAND TOTAL at end.
    Phase sheet columns (col 2-8 since col 1 is empty gutter):
      Date | Day | Timing | Hours | Files | Evidence/Theme | What I Did
    """
    if args.phase not in wb.sheetnames:
        sys.stderr.write(f"ERROR: sheet '{args.phase}' missing from workbook\n")
        sys.exit(4)
    ws = wb[args.phase]

    gt_row = _find_grand_total_row(ws)
    if gt_row is None:
        sys.stderr.write(f"ERROR: GRAND TOTAL row not found in '{args.phase}'\n")
        sys.exit(5)

    target_row = gt_row
    ws.insert_rows(gt_row)
    new_gt_row = gt_row + 1

    timing = f"{args.start} – {args.end}"
    # cols B..H (2..8) — col offset +1 for gutter
    ws.cell(row=target_row, column=1 + COL_OFFSET, value=datetime.strptime(args.date, "%Y-%m-%d")).number_format = "yyyy-mm-dd"
    ws.cell(row=target_row, column=2 + COL_OFFSET, value=args.day)
    ws.cell(row=target_row, column=3 + COL_OFFSET, value=timing)
    ws.cell(row=target_row, column=4 + COL_OFFSET, value=args.hours).number_format = "0.00"
    ws.cell(row=target_row, column=5 + COL_OFFSET, value=args.files)
    ws.cell(row=target_row, column=6 + COL_OFFSET, value=args.theme)
    ws.cell(row=target_row, column=7 + COL_OFFSET, value=args.what)

    # Apply consistent styling on cols B..H (skip col A gutter)
    for c in range(1 + COL_OFFSET, 8 + COL_OFFSET):
        cell = ws.cell(row=target_row, column=c)
        cell.font = NORM_FONT
        cell.alignment = CTR if c in (1 + COL_OFFSET, 2 + COL_OFFSET, 4 + COL_OFFSET, 5 + COL_OFFSET) else WRAP
        cell.border = BR

    ws.row_dimensions[target_row].height = 64

    # Header is at row 5 in NEW layout
    _extend_sum_formulas(ws, header_row=5, gt_row_after_insert=new_gt_row)

    return target_row


def append_to_chrono_sheet(wb, args: argparse.Namespace) -> int:
    """Append the row to 'All Sessions'. Returns the new row index.

    Day-11 layout: sheet renamed from 'All Sessions (chronological)' → 'All Sessions'.
    Header at row 5; data row 6+; GRAND TOTAL at end.
    Columns (col 2-10 since col 1 is gutter):
      Date | Day | Timing | Hours | Files | Phase | Evidence/Theme | What I Did | Parallel Work (FE+BE)
    Parallel Work column (J) is populated by the correlation hook (Day-11 TASK 6) — left blank by this script.
    """
    sheet_name = "All Sessions"
    if sheet_name not in wb.sheetnames:
        # Fallback for any worktree that hasn't pulled the Day-11 swap yet
        legacy = "All Sessions (chronological)"
        if legacy in wb.sheetnames:
            sheet_name = legacy
        else:
            sys.stderr.write(f"ERROR: neither 'All Sessions' nor '{legacy}' sheet found\n")
            sys.exit(4)
    ws = wb[sheet_name]

    gt_row = _find_grand_total_row(ws)
    if gt_row is None:
        sys.stderr.write(f"ERROR: GRAND TOTAL row not found in '{sheet_name}'\n")
        sys.exit(5)

    target_row = gt_row
    ws.insert_rows(gt_row)
    new_gt_row = gt_row + 1

    timing = f"{args.start} – {args.end}"
    # cols B..I (2..9) — col offset +1 for gutter; col J left blank for correlation hook
    ws.cell(row=target_row, column=1 + COL_OFFSET, value=datetime.strptime(args.date, "%Y-%m-%d")).number_format = "yyyy-mm-dd"
    ws.cell(row=target_row, column=2 + COL_OFFSET, value=args.day)
    ws.cell(row=target_row, column=3 + COL_OFFSET, value=timing)
    ws.cell(row=target_row, column=4 + COL_OFFSET, value=args.hours).number_format = "0.00"
    ws.cell(row=target_row, column=5 + COL_OFFSET, value=args.files)
    ws.cell(row=target_row, column=6 + COL_OFFSET, value=args.phase)
    ws.cell(row=target_row, column=7 + COL_OFFSET, value=args.theme)
    ws.cell(row=target_row, column=8 + COL_OFFSET, value=args.what)
    # col 9+1 = col 10 (J) "Parallel Work (FE+BE)" — left blank; populated by Day-11 TASK 6 hook later

    for c in range(1 + COL_OFFSET, 10 + COL_OFFSET):
        cell = ws.cell(row=target_row, column=c)
        cell.font = NORM_FONT
        cell.alignment = CTR if c in (1 + COL_OFFSET, 2 + COL_OFFSET, 4 + COL_OFFSET, 5 + COL_OFFSET) else WRAP
        cell.border = BR

    ws.row_dimensions[target_row].height = 64

    _extend_sum_formulas(ws, header_row=5, gt_row_after_insert=new_gt_row)

    return target_row


def main() -> int:
    args = parse_args()
    validate(args)

    print(f"Loading workbook: {WORKBOOK_PATH}")
    wb = load_workbook(WORKBOOK_PATH)

    print(f"Appending row to phase sheet: '{args.phase}'")
    phase_row = append_to_phase_sheet(wb, args)

    print(f"Appending row to chronological sheet: 'All Sessions'")
    chrono_row = append_to_chrono_sheet(wb, args)

    print()
    print("Row to be written:")
    print(f"  date    {args.date} ({args.day})")
    print(f"  timing  {args.start} – {args.end}")
    print(f"  hours   {args.hours}")
    print(f"  files   {args.files}")
    print(f"  phase   {args.phase}")
    print(f"  theme   {args.theme}")
    print(f"  what    {args.what[:120]}{'...' if len(args.what) > 120 else ''}")
    print()
    print(f"Phase sheet '{args.phase}' row: {phase_row}")
    print(f"Chrono sheet row: {chrono_row}")

    if args.dry_run:
        print()
        print("DRY-RUN: workbook NOT saved.")
        return 0

    wb.save(WORKBOOK_PATH)
    print()
    print(f"✓ Saved {WORKBOOK_PATH}")
    print()
    print("NOTE (Day-11 layout): GRAND TOTAL =SUM(...) formulas auto-extended to cover the new row.")
    print("      Excel still re-evaluates Day-Total subtotals on open if any are present in the sheet.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
