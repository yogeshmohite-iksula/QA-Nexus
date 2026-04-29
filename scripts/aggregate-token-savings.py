#!/usr/bin/env python3
"""
Aggregate token-savings across all 3 worktrees (MAIN + FE + BE) into a
3-sheet Excel at docs/observability/Token_Savings_Log.xlsx.

Sheets:
  1. Sessions         — chronological, one row per session JSONL line
  2. Daily Rollup     — grouped by date, sum across all chats
  3. Per-Chat Compare — MAIN vs FE vs BE totals + tokens-per-hour

Reads:
  ./.claude/token-savings.jsonl                     (MAIN worktree)
  ../Project10-QA_Nexus-frontend/.claude/token-savings.jsonl  (FE, optional)
  ../Project10-QA_Nexus-backend/.claude/token-savings.jsonl   (BE, optional)

Writes:
  docs/observability/Token_Savings_Log.xlsx

Run via:
  python3 scripts/aggregate-token-savings.py

Or via the /eod-tokens slash command which also prints a summary table.

Styling matches QA_Nexus_Work_Hours_Log.xlsx convention:
  - teal header row (#0F766E) with white bold text
  - frozen header pane
  - daily subtotal rows
  - grand total row at bottom of Sessions sheet
"""
from __future__ import annotations

import json
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print(
        "ERROR: openpyxl is required.\n"
        "  pip3 install --user openpyxl\n"
        "or run via: pnpm dlx --package openpyxl python3 scripts/aggregate-token-savings.py",
        file=sys.stderr,
    )
    sys.exit(1)


# ─── Configuration ────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent.parent
WORKTREE_PATHS = {
    "MAIN": ROOT / ".claude" / "token-savings.jsonl",
    "FE": ROOT.parent / "Project10-QA_Nexus-frontend" / ".claude" / "token-savings.jsonl",
    "BE": ROOT.parent / "Project10-QA_Nexus-backend" / ".claude" / "token-savings.jsonl",
}
OUTPUT = ROOT / "docs" / "observability" / "Token_Savings_Log.xlsx"

# Styling constants — match QA_Nexus_Work_Hours_Log.xlsx
TEAL_HEADER = PatternFill("solid", fgColor="0F766E")
LIGHT_GRAY = PatternFill("solid", fgColor="F1F5F9")
SUBTOTAL_FILL = PatternFill("solid", fgColor="E5E7EB")
GRAND_TOTAL_FILL = PatternFill("solid", fgColor="2DD4BF")  # teal accent
WHITE_BOLD = Font(name="Inter", size=11, bold=True, color="FFFFFF")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(horizontal="left", vertical="center")
NUMBER_ALIGN = Alignment(horizontal="right", vertical="center")
THIN = Side(border_style="thin", color="C7D0DC")
BORDER = Border(top=THIN, left=THIN, right=THIN, bottom=THIN)


def load_jsonl(path: Path, default_role: str | None = None) -> list[dict]:
    """Load all lines from a JSONL file. Returns [] if file missing."""
    if not path.exists():
        return []
    rows: list[dict] = []
    with path.open() as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
                if default_role and not row.get("chat_role"):
                    row["chat_role"] = default_role
                rows.append(row)
            except json.JSONDecodeError:
                # malformed line — skip but warn
                print(f"  WARN: malformed JSON in {path}: {line[:80]}", file=sys.stderr)
    return rows


def load_all_sessions() -> list[dict]:
    """Walk all 3 worktree JSONL files, dedup by session, return merged list.

    DEDUP CONTRACT: each Stop event re-snapshots the FULL running session state
    (cumulative tool_calls, memory_injects, etc — not deltas). So a single
    Claude session with 5 Stop fires writes 5 rows to the JSONL where row N+1
    is row N's superset. SUM-aggregating those rows over-counts by 5×.

    The fix: dedup by (session_id, chat_role) and KEEP THE LAST ROW per
    session (the highest-watermark snapshot). After dedup, daily SUM is
    correct because each session contributes exactly once with its full
    final-state numbers.

    "Last" is determined by ended_at descending; falls back to started_at
    then to file-order if both missing. session_id "" is treated as a
    separate bucket per chat_role so legacy malformed rows don't collapse.

    See `docs/observability/token-tracking.md` § "Estimation rules" for
    the rationale + Day-3 EOD report § "Token-savings dashboard" for the
    incident that motivated this dedup (Day-2 number ballooned 187K → 974K
    when overnight sessions accumulated multiple Stop fires).
    """
    all_sessions: list[dict] = []
    for role, path in WORKTREE_PATHS.items():
        rows = load_jsonl(path, default_role=role)
        if rows:
            print(f"  ✓ {role}: {len(rows)} raw rows from {path}")
        else:
            print(f"  · {role}: no data (skipped)")
        all_sessions.extend(rows)

    # ─── dedup by (session_id, chat_role) — keep latest snapshot ─────
    # Sort ascending by ended_at so the LAST entry wins when we walk dict.
    def _sort_key(r: dict) -> tuple[str, str, str]:
        return (
            r.get("ended_at", "") or "",
            r.get("started_at", "") or "",
            r.get("date", "") or "",
        )

    all_sessions.sort(key=_sort_key)
    deduped_by_key: dict[tuple[str, str], dict] = {}
    for r in all_sessions:
        key = (r.get("session_id", "") or "", r.get("chat_role", "") or "")
        deduped_by_key[key] = r  # later key replaces earlier (keep latest snapshot)

    deduped = list(deduped_by_key.values())
    duplicates_dropped = len(all_sessions) - len(deduped)
    if duplicates_dropped > 0:
        print(
            f"  ↓ deduped {duplicates_dropped} stale Stop-snapshot rows "
            f"({len(all_sessions)} → {len(deduped)} sessions)"
        )

    # Sort chronologically for downstream consumers (chronological sheet, etc).
    deduped.sort(
        key=lambda r: (r.get("date", ""), r.get("ended_at", r.get("started_at", "")))
    )
    return deduped


def style_header(cell) -> None:
    cell.fill = TEAL_HEADER
    cell.font = WHITE_BOLD
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER


def style_data(cell, alt: bool = False, numeric: bool = False) -> None:
    cell.font = Font(name="Inter", size=10)
    cell.alignment = NUMBER_ALIGN if numeric else DATA_ALIGN
    cell.border = BORDER
    if alt:
        cell.fill = LIGHT_GRAY


def style_subtotal(cell) -> None:
    cell.fill = SUBTOTAL_FILL
    cell.font = Font(name="Inter", size=10, bold=True)
    cell.alignment = NUMBER_ALIGN
    cell.border = BORDER


def style_grand_total(cell) -> None:
    cell.fill = GRAND_TOTAL_FILL
    cell.font = Font(name="Inter", size=11, bold=True, color="0B0F17")
    cell.alignment = NUMBER_ALIGN
    cell.border = BORDER


def build_sessions_sheet(ws, sessions: list[dict]) -> None:
    """Sheet 1: chronological session log with daily subtotals."""
    headers = [
        "Date",
        "Chat",
        "Started",
        "Ended",
        "Duration (min)",
        "Tool Calls",
        "Memory Injects",
        "Context Preloads",
        "Tokens Saved",
        "Branch",
        "Commits",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        style_header(cell)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Group sessions by date for subtotals
    rows_by_date: dict[str, list[dict]] = defaultdict(list)
    for s in sessions:
        rows_by_date[s.get("date", "—")].append(s)

    row_idx = 2
    grand_totals = {"duration_min": 0, "tool_calls": 0, "memory_injects": 0,
                     "context_preloads": 0, "tokens_saved": 0, "commits": 0}

    for date in sorted(rows_by_date.keys()):
        sessions_today = rows_by_date[date]
        for i, s in enumerate(sessions_today):
            alt = i % 2 == 1
            values = [
                s.get("date", ""),
                s.get("chat_role", ""),
                s.get("started_at", "").split("T")[1].rstrip("Z") if "T" in s.get("started_at", "") else "",
                s.get("ended_at", "").split("T")[1].rstrip("Z") if "T" in s.get("ended_at", "") else "",
                s.get("duration_min", 0),
                s.get("tool_calls", 0),
                s.get("memory_injects", 0),
                s.get("context_preloads", 0),
                s.get("tokens_saved_estimated", 0),
                s.get("branch", ""),
                s.get("commits", 0),
            ]
            for col, v in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=v)
                style_data(cell, alt=alt, numeric=col in (5, 6, 7, 8, 9, 11))
            row_idx += 1

        # Daily subtotal row
        subtotal = {
            "duration_min": sum(s.get("duration_min", 0) for s in sessions_today),
            "tool_calls": sum(s.get("tool_calls", 0) for s in sessions_today),
            "memory_injects": sum(s.get("memory_injects", 0) for s in sessions_today),
            "context_preloads": sum(s.get("context_preloads", 0) for s in sessions_today),
            "tokens_saved": sum(s.get("tokens_saved_estimated", 0) for s in sessions_today),
            "commits": sum(s.get("commits", 0) for s in sessions_today),
        }
        sub_values = [
            f"{date} TOTAL", "", "", "",
            subtotal["duration_min"], subtotal["tool_calls"],
            subtotal["memory_injects"], subtotal["context_preloads"],
            subtotal["tokens_saved"], "", subtotal["commits"],
        ]
        for col, v in enumerate(sub_values, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            style_subtotal(cell)
        row_idx += 1
        for k, v in subtotal.items():
            grand_totals[k] = grand_totals.get(k, 0) + v

    # Grand total at the bottom
    grand_values = [
        "GRAND TOTAL", "", "", "",
        grand_totals["duration_min"], grand_totals["tool_calls"],
        grand_totals["memory_injects"], grand_totals["context_preloads"],
        grand_totals["tokens_saved"], "", grand_totals["commits"],
    ]
    for col, v in enumerate(grand_values, 1):
        cell = ws.cell(row=row_idx, column=col, value=v)
        style_grand_total(cell)

    # Column widths — readable
    widths = [12, 6, 10, 10, 14, 12, 14, 16, 14, 35, 9]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_daily_rollup_sheet(ws, sessions: list[dict]) -> None:
    """Sheet 2: one row per date — total sessions, hours, tokens."""
    headers = ["Date", "Total Sessions", "Total Hours", "Total Tool Calls", "Total Tokens Saved"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        style_header(cell)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    by_date: dict[str, dict] = defaultdict(lambda: {"sessions": 0, "minutes": 0, "tools": 0, "tokens": 0})
    for s in sessions:
        d = by_date[s.get("date", "—")]
        d["sessions"] += 1
        d["minutes"] += s.get("duration_min", 0)
        d["tools"] += s.get("tool_calls", 0)
        d["tokens"] += s.get("tokens_saved_estimated", 0)

    row_idx = 2
    for date in sorted(by_date.keys()):
        d = by_date[date]
        values = [date, d["sessions"], round(d["minutes"] / 60, 2), d["tools"], d["tokens"]]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            style_data(cell, alt=(row_idx % 2 == 0), numeric=col in (2, 3, 4, 5))
        row_idx += 1

    # Add SUM formulas at bottom
    total_row = row_idx
    ws.cell(row=total_row, column=1, value="GRAND TOTAL")
    style_grand_total(ws.cell(row=total_row, column=1))
    for col in (2, 3, 4, 5):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col, value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})")
        style_grand_total(cell)

    widths = [14, 16, 14, 18, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_per_chat_sheet(ws, sessions: list[dict]) -> None:
    """Sheet 3: MAIN vs FE vs BE comparison."""
    headers = ["Chat", "Sessions", "Hours", "Tokens Saved", "Tokens / Hour"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        style_header(cell)
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    by_chat: dict[str, dict] = defaultdict(lambda: {"sessions": 0, "minutes": 0, "tokens": 0})
    for s in sessions:
        c = by_chat[s.get("chat_role", "?")]
        c["sessions"] += 1
        c["minutes"] += s.get("duration_min", 0)
        c["tokens"] += s.get("tokens_saved_estimated", 0)

    row_idx = 2
    for chat in ("MAIN", "FE", "BE"):
        if chat not in by_chat:
            continue
        c = by_chat[chat]
        hours = round(c["minutes"] / 60, 2)
        tph = round(c["tokens"] / hours) if hours > 0 else 0
        values = [chat, c["sessions"], hours, c["tokens"], tph]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=v)
            style_data(cell, alt=(row_idx % 2 == 0), numeric=col in (2, 3, 4, 5))
        row_idx += 1

    # Grand total
    total_row = row_idx
    ws.cell(row=total_row, column=1, value="GRAND TOTAL")
    style_grand_total(ws.cell(row=total_row, column=1))
    for col in (2, 3, 4):
        col_letter = get_column_letter(col)
        cell = ws.cell(row=total_row, column=col, value=f"=SUM({col_letter}2:{col_letter}{total_row - 1})")
        style_grand_total(cell)
    # Tokens / Hour grand = D / C
    cell = ws.cell(row=total_row, column=5, value=f"=IFERROR(D{total_row}/C{total_row},0)")
    style_grand_total(cell)

    widths = [12, 12, 10, 16, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def main() -> int:
    print(f"Aggregating token-savings across worktrees → {OUTPUT}")
    sessions = load_all_sessions()
    if not sessions:
        print("\nNo session data found. Have any Stop hooks fired yet?", file=sys.stderr)
        return 1

    OUTPUT.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet
    ws1 = wb.create_sheet("Sessions")
    build_sessions_sheet(ws1, sessions)
    ws2 = wb.create_sheet("Daily Rollup")
    build_daily_rollup_sheet(ws2, sessions)
    ws3 = wb.create_sheet("Per-Chat Comparison")
    build_per_chat_sheet(ws3, sessions)

    wb.save(OUTPUT)

    # Print summary table to stdout
    print(f"\n✓ Wrote {len(sessions)} sessions to {OUTPUT.relative_to(ROOT)}\n")
    print("Daily rollup:")
    by_date: dict[str, int] = defaultdict(int)
    by_date_min: dict[str, int] = defaultdict(int)
    for s in sessions:
        by_date[s.get("date", "—")] += s.get("tokens_saved_estimated", 0)
        by_date_min[s.get("date", "—")] += s.get("duration_min", 0)
    for d in sorted(by_date.keys()):
        h = round(by_date_min[d] / 60, 1)
        print(f"  {d}  {h:>5} hr   {by_date[d]:>8,} tokens saved")

    return 0


if __name__ == "__main__":
    sys.exit(main())
