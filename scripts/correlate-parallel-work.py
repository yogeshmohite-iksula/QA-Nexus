#!/usr/bin/env python3
"""
correlate-parallel-work.py — populate the "Parallel Work (FE+BE)" column
in docs/observability/QA_Nexus_Work_Log.xlsx All Sessions sheet.

Spec: Day-11 addendum TASK 6 / Day-12 TASK 3 — `feat(observability):
correlate FE+BE work into MAIN session rows`.

Reads docs/observability/sessions-stream.jsonl (all sessions across all
worktrees, appended by .claude/hooks/stop/log-session-summary.sh per
session end) and for each MAIN row in All Sessions for a given date,
finds FE/BE entries with > 50% time overlap and writes a combined
summary into column J.

Design choices:
- 50% overlap threshold = max(0, min(end1, end2) - max(start1, start2))
                         / min(duration1, duration2)
- Multiple matches per worktree → take the max-overlap one
- Output format: "FE: <FE summary>; BE: <BE summary>" (drop one if absent)
- Truncate per-worktree summary to ~200 chars to stay readable in Excel
- Idempotent: overwrites existing column J value for the date

Usage:
    python3 scripts/correlate-parallel-work.py --date 2026-05-07
    python3 scripts/correlate-parallel-work.py --date 2026-05-07 --dry-run

Cross-references:
- .claude/hooks/stop/log-session-summary.sh — writes the JSONL stream
- scripts/update-work-log.py — appends new rows (paired tool)
- docs/observability/README.md — convention overview
"""

from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime, date as date_cls
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.stderr.write("ERROR: openpyxl not installed. Run: pip install openpyxl --break-system-packages\n")
    sys.exit(1)

THIS = Path(__file__).resolve()
REPO_ROOT = THIS.parent.parent
WORKBOOK_PATH = REPO_ROOT / "docs" / "observability" / "QA_Nexus_Work_Log.xlsx"
STREAM_PATH = REPO_ROOT / "docs" / "observability" / "sessions-stream.jsonl"

# All Sessions sheet layout (Day-11 Professional layout):
#   Col A: empty gutter
#   Col B: Date     Col C: Day     Col D: Timing
#   Col E: Hours    Col F: Files   Col G: Phase
#   Col H: Evidence/Theme   Col I: What I Did   Col J: Parallel Work (FE+BE)
HEADER_ROW = 5
COL_DATE = 2
COL_TIMING = 4
COL_PARALLEL = 10  # J
SHEET_NAME = "All Sessions"

# Per-worktree summary char cap (keeps Excel cell readable).
SUMMARY_CAP = 200
# Combined cell content cap — Excel handles >32k but we keep it neat.
CELL_CAP = 500


def _parse_hhmm(s: str) -> int | None:
    """Convert 'HH:MM' or 'H:MM AM/PM' to minutes-since-midnight; None on failure."""
    if not s:
        return None
    s = s.strip()
    for fmt in ("%H:%M", "%I:%M %p", "%I:%M%p", "%I %p"):
        try:
            t = datetime.strptime(s, fmt)
            return t.hour * 60 + t.minute
        except ValueError:
            continue
    return None


def _parse_timing_range(timing: str) -> tuple[int, int] | None:
    """Parse 'H:MM AM – H:MM PM' (with en-dash or hyphen) to (start_min, end_min)."""
    if not timing:
        return None
    for sep in (" – ", " - ", "–", "-"):
        if sep in timing:
            parts = timing.split(sep, 1)
            if len(parts) == 2:
                a = _parse_hhmm(parts[0])
                b = _parse_hhmm(parts[1])
                if a is not None and b is not None:
                    return (min(a, b), max(a, b))
    return None


def _overlap_minutes(a: tuple[int, int], b: tuple[int, int]) -> int:
    """Time-overlap in minutes between two (start, end) ranges."""
    return max(0, min(a[1], b[1]) - max(a[0], b[0]))


def _norm_date(v) -> str:
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date_cls):
        return v.isoformat()
    return str(v).strip() if v else ""


def load_sessions_for_date(target_date: str) -> list[dict]:
    """Read sessions-stream.jsonl, return rows matching target_date."""
    if not STREAM_PATH.exists():
        return []
    out = []
    for line in STREAM_PATH.read_text().splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            rec = json.loads(line)
        except json.JSONDecodeError:
            continue
        if rec.get("date") == target_date:
            out.append(rec)
    return out


def correlate(target_date: str, dry_run: bool = False) -> int:
    """Update All Sessions col J for MAIN rows on target_date.

    Returns count of rows updated.
    """
    if not WORKBOOK_PATH.exists():
        sys.stderr.write(f"ERROR: workbook not found at {WORKBOOK_PATH}\n")
        sys.exit(3)

    sessions = load_sessions_for_date(target_date)
    if not sessions:
        print(f"No sessions in stream for {target_date}; nothing to correlate.")
        return 0

    fe_rows = [s for s in sessions if s.get("worktree") == "FE"]
    be_rows = [s for s in sessions if s.get("worktree") == "BE"]
    main_rows = [s for s in sessions if s.get("worktree") == "MAIN"]

    print(f"Stream for {target_date}: {len(main_rows)} MAIN, {len(fe_rows)} FE, {len(be_rows)} BE")

    wb = load_workbook(WORKBOOK_PATH, data_only=False)
    if SHEET_NAME not in wb.sheetnames:
        sys.stderr.write(f"ERROR: '{SHEET_NAME}' sheet missing\n")
        sys.exit(4)
    ws = wb[SHEET_NAME]

    updates = 0

    # Iterate All Sessions rows for the target date. We only populate col J
    # for rows that look like MAIN (Phase = "Milestone Planning" is the
    # heuristic; falls back to populating regardless if uncertain).
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        d = _norm_date(ws.cell(row=r, column=COL_DATE).value)
        if d != target_date:
            continue
        timing = ws.cell(row=r, column=COL_TIMING).value
        rng = _parse_timing_range(timing) if isinstance(timing, str) else None
        if rng is None:
            continue

        # Find best FE match
        fe_best = _best_overlap(fe_rows, rng)
        be_best = _best_overlap(be_rows, rng)

        parts = []
        if fe_best:
            parts.append(f"FE: {(fe_best.get('summary') or '')[:SUMMARY_CAP]}")
        if be_best:
            parts.append(f"BE: {(be_best.get('summary') or '')[:SUMMARY_CAP]}")
        if not parts:
            continue

        cell_value = "; ".join(parts)[:CELL_CAP]
        existing = ws.cell(row=r, column=COL_PARALLEL).value
        if existing == cell_value:
            continue
        if dry_run:
            print(f"  WOULD UPDATE R{r} (timing={timing}):\n    {cell_value[:120]}{'...' if len(cell_value) > 120 else ''}")
        else:
            ws.cell(row=r, column=COL_PARALLEL).value = cell_value
            print(f"  UPDATED R{r} (timing={timing}): {len(parts)} parallel-tracks correlated")
        updates += 1

    if not dry_run and updates > 0:
        wb.save(WORKBOOK_PATH)
        print(f"\n✓ Saved {WORKBOOK_PATH} ({updates} row(s) updated)")
    elif dry_run:
        print(f"\nDRY-RUN: would update {updates} row(s); workbook NOT saved.")
    else:
        print("\nNo updates needed (already correlated or no overlap).")
    return updates


def _best_overlap(rows: list[dict], target_range: tuple[int, int]) -> dict | None:
    """Pick the row with > 50% overlap and the largest absolute overlap.
    Threshold: overlap / min(duration_target, duration_row) > 0.5.
    """
    target_dur = target_range[1] - target_range[0]
    if target_dur <= 0:
        return None
    best = None
    best_overlap = 0
    for r in rows:
        s = _parse_hhmm(r.get("start", ""))
        e = _parse_hhmm(r.get("end", ""))
        if s is None or e is None or e <= s:
            continue
        ov = _overlap_minutes(target_range, (s, e))
        if ov == 0:
            continue
        denom = min(target_dur, e - s)
        if ov / denom <= 0.5:
            continue
        if ov > best_overlap:
            best_overlap = ov
            best = r
    return best


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Correlate FE+BE sessions into MAIN row column J in the work-log xlsx.")
    p.add_argument("--date", required=True, help="Target date YYYY-MM-DD. Correlates rows on this date.")
    p.add_argument("--dry-run", action="store_true", help="Print what would change; do not save.")
    return p.parse_args()


def main() -> int:
    args = parse_args()
    try:
        datetime.strptime(args.date, "%Y-%m-%d")
    except ValueError:
        sys.stderr.write(f"ERROR: --date must be YYYY-MM-DD, got '{args.date}'\n")
        return 2
    correlate(args.date, dry_run=args.dry_run)
    return 0


if __name__ == "__main__":
    sys.exit(main())
