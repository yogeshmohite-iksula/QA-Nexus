#!/usr/bin/env python3
"""
Rebuild the `Tokens —` sheets in QA_Nexus_Work_Log.xlsx from raw sources,
adding Tokens In / Tokens Out / Cost USD per session via Option β
(read ~/.claude/projects/...<sessionid>.jsonl for actual usage data) with
Option α (heuristic) fallback when conversation file missing/unparseable.

Sources:
  1. .claude/token-savings.jsonl (MAIN worktree) — per-session Stop-hook log
  2. ~/.claude/projects/<project>/<session_id>.jsonl — Claude Code's per-message log
  3. (FE/BE worktrees: jsonl files don't exist yet; rows for those = <unknown>)

Writes 4 sheets to docs/observability/QA_Nexus_Work_Log.xlsx:
  - `Tokens — Sessions` (REPLACE; full chronological per-session)
  - `Tokens — Daily Rollup` (REPLACE)
  - `Tokens — Per-Chat Comparison` (REPLACE)
  - `Backfill_Notes` (NEW; data-quality flags)

Cost model (Sonnet 4.7, May 2026 pricing):
  - Input: $3.00 / Mtok
  - Output: $15.00 / Mtok
  - Cache read: $0.30 / Mtok
  - Cache write (creation): $3.75 / Mtok (1.25× input)

Source column tagging:
  - `conversation` = Option β (real token counts from conversation file)
  - `heuristic` = Option α (estimated; tokens_in ≈ tool_calls × 2000)
  - `manual` = operator-entered row
  - `jsonl-only` = Stop-hook log only (savings present, tokens used unknown)

Run: python3 scripts/rebuild-work-log-tokens.py
"""
from __future__ import annotations

import json
import os
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    print("ERROR: openpyxl required. pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)

REPO_ROOT = Path(__file__).resolve().parent.parent
WORK_LOG = REPO_ROOT / "docs" / "observability" / "QA_Nexus_Work_Log.xlsx"

# Pricing (Sonnet 4.7, USD/Mtok)
RATE_IN = 3.00
RATE_OUT = 15.00
RATE_CACHE_READ = 0.30
RATE_CACHE_WRITE = 3.75

# Day-N anchor
DAY_0 = datetime(2026, 4, 26).date()

# Where Claude Code stores per-session conversation files
CONV_DIR = Path.home() / ".claude" / "projects" / "-Users-yogeshmohite-AI-Tester-Project-Project10-QA-Nexus"


def day_n(date_str: str) -> int:
    """YYYY-MM-DD → Day-N integer."""
    d = datetime.strptime(date_str, "%Y-%m-%d").date()
    return (d - DAY_0).days


def conversation_tokens(session_id: str) -> dict | None:
    """Option β: read the per-session conversation file + sum all assistant
    message usage. Returns dict with totals or None if file missing/unparseable."""
    if not session_id:
        return None
    fpath = CONV_DIR / f"{session_id}.jsonl"
    if not fpath.exists():
        return None
    totals = {"input": 0, "output": 0, "cache_read": 0, "cache_write": 0}
    try:
        with fpath.open() as f:
            for line in f:
                try:
                    obj = json.loads(line)
                except (json.JSONDecodeError, ValueError):
                    continue
                msg = obj.get("message") or {}
                usage = msg.get("usage") or obj.get("usage")
                if not usage or not isinstance(usage, dict):
                    continue
                totals["input"] += usage.get("input_tokens", 0) or 0
                totals["output"] += usage.get("output_tokens", 0) or 0
                totals["cache_read"] += usage.get("cache_read_input_tokens", 0) or 0
                totals["cache_write"] += usage.get("cache_creation_input_tokens", 0) or 0
    except (IOError, OSError):
        return None
    if totals["input"] == 0 and totals["output"] == 0 and totals["cache_read"] == 0:
        return None
    return totals


def heuristic_tokens(tool_calls: int) -> dict:
    """Option α: rough estimate when conversation file unavailable."""
    return {
        "input": tool_calls * 2000,
        "output": tool_calls * 500,
        "cache_read": 0,
        "cache_write": 0,
    }


def compute_cost(tokens: dict) -> float:
    """USD cost from token totals."""
    return round(
        (tokens["input"] * RATE_IN + tokens["output"] * RATE_OUT + tokens["cache_read"] * RATE_CACHE_READ + tokens["cache_write"] * RATE_CACHE_WRITE) / 1_000_000,
        4,
    )


def load_sessions() -> tuple[list[dict], list[dict]]:
    """Load all sessions from .claude/token-savings.jsonl (MAIN worktree).
    Returns (sessions, backfill_notes)."""
    jsonl_path = REPO_ROOT / ".claude" / "token-savings.jsonl"
    if not jsonl_path.exists():
        return [], [{"date": "n/a", "worktree": "MAIN", "field": "all", "reason": "token-savings.jsonl missing", "mitigation": "verify Stop hook is firing"}]

    # First pass: load all rows, dedupe by session_id keeping the MAX tool_calls
    # snapshot (Stop hook fires multiple times per session as agent restarts;
    # each fire is a cumulative snapshot, so the highest tool_calls is the
    # final state). This mirrors the existing aggregator's dedupe behavior
    # ("deduped 79 stale Stop-snapshot rows (86 → 7 sessions)").
    raw_rows = []
    with jsonl_path.open() as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                obj = json.loads(line)
            except (json.JSONDecodeError, ValueError):
                continue
            raw_rows.append(obj)

    # Group by session_id; keep the row with max tool_calls per group
    by_sid: dict[str, dict] = {}
    for obj in raw_rows:
        sid = obj.get("session_id", "")
        if not sid:
            continue
        existing = by_sid.get(sid)
        if existing is None or (obj.get("tool_calls", 0) or 0) > (existing.get("tool_calls", 0) or 0):
            by_sid[sid] = obj

    sessions = []
    notes = []

    for obj in by_sid.values():

            # Try Option β first
            sid = obj.get("session_id", "")
            tokens = conversation_tokens(sid)
            source = "conversation"
            if tokens is None:
                tokens = heuristic_tokens(obj.get("tool_calls", 0))
                source = "heuristic"
                notes.append({
                    "date": obj.get("date", "?"),
                    "worktree": obj.get("chat_role", "MAIN"),
                    "field": "Tokens In/Out",
                    "reason": f"Conversation file ~/.claude/projects/.../{sid[:8]}.jsonl missing or empty",
                    "mitigation": "Heuristic estimate (tool_calls × 2000 in / × 500 out); refine when conversation reachable",
                })

            cost_usd = compute_cost(tokens)
            in_out_ratio = round(tokens["input"] / tokens["output"], 2) if tokens["output"] else 0.0

            sessions.append({
                "date": obj.get("date", ""),
                "day_n": day_n(obj.get("date", "")) if obj.get("date") else None,
                "worktree": obj.get("chat_role", "MAIN"),
                "started": obj.get("started_at", "")[11:16],
                "ended": obj.get("ended_at", "")[11:16],
                "duration_min": obj.get("duration_min", 0),
                "tool_calls": obj.get("tool_calls", 0),
                "tokens_in": tokens["input"] + tokens["cache_write"],  # cache_write billed as input
                "tokens_out": tokens["output"],
                "tokens_cache_read": tokens["cache_read"],
                "tokens_saved": obj.get("tokens_saved_estimated", 0),
                "cost_usd": cost_usd,
                "in_out_ratio": in_out_ratio,
                "branch": obj.get("branch", ""),
                "commits": obj.get("commits", 0),
                "source": source,
                "session_id": sid,
            })

    sessions.sort(key=lambda r: (r["date"], r["started"]))
    return sessions, notes


def add_fe_be_placeholder_rows(sessions: list[dict], notes: list[dict]) -> tuple[list[dict], list[dict]]:
    """Ensure each (Day-N, worktree) combo for Days 0-7 has a row. Missing
    days get <unknown> placeholders + a Backfill_Notes entry. Plus: if MAIN
    has no data on a day, add MAIN <unknown> too (Stop hook didn't fire that
    day; reconstruct from EOD report if needed)."""
    from datetime import timedelta
    # Generate Day 0 → Day 7 dates
    target_dates = [(DAY_0 + timedelta(days=n)).isoformat() for n in range(8)]
    for date in target_dates:
        for wt in ("MAIN", "FE", "BE"):
            already_has = any(s["worktree"] == wt and s["date"] == date for s in sessions)
            if already_has:
                continue
            sessions.append({
                "date": date,
                "day_n": day_n(date),
                "worktree": wt,
                "started": "<unknown>",
                "ended": "<unknown>",
                "duration_min": 0,
                "tool_calls": 0,
                "tokens_in": 0,
                "tokens_out": 0,
                "tokens_cache_read": 0,
                "tokens_saved": 0,
                "cost_usd": 0.0,
                "in_out_ratio": 0.0,
                "branch": "",
                "commits": 0,
                "source": "<unknown>",
                "session_id": "",
            })
            if wt == "MAIN":
                reason = f"Stop hook didn't write a token-savings.jsonl row on {date} (likely session cleanly closed without Stop event, OR session crossed midnight UTC)"
                mitigation = f"Reconstruct hours from docs/eod-reports/{date}-day-{day_n(date)}.md; tokens unrecoverable for past sessions"
            elif wt == "FE":
                reason = "FE worktree Stop hook NEVER fired (audit.jsonl absent — Bug A in token-tracking-diagnostic-2026-05-04.md)"
                mitigation = "Operator self-report from FE+1 EOD summary; followup (z) to investigate hook activation Day 9 morning"
            else:  # BE
                reason = "BE worktree Stop hook fired through 2026-04-28 then stopped (Bug B in diagnostic doc)"
                mitigation = "Operator self-report from BE+1 EOD summary; same followup (z) as FE"
            notes.append({
                "date": date,
                "worktree": wt,
                "field": "All token + hours fields",
                "reason": reason,
                "mitigation": mitigation,
            })
    sessions.sort(key=lambda r: (r["date"], r["worktree"], r["started"]))
    return sessions, notes


def write_workbook(sessions: list[dict], notes: list[dict]) -> None:
    """Replace 3 Tokens sheets + add/refresh Backfill_Notes in the work-log."""
    wb = load_workbook(WORK_LOG)
    HEADER_FILL = PatternFill(start_color="0F766E", end_color="0F766E", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF")

    def reset_sheet(name: str, headers: list[str]):
        if name in wb.sheetnames:
            del wb[name]
        ws = wb.create_sheet(name)
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"
        return ws

    # 1. Tokens — Sessions
    ws_s = reset_sheet("Tokens — Sessions", [
        "Date", "Day-N", "Worktree", "Started", "Ended", "Duration (min)",
        "Tool Calls", "Tokens In", "Tokens Out", "Cache Read", "Tokens Saved",
        "Cost USD", "In:Out Ratio", "Branch", "Commits", "Source",
    ])
    for s in sessions:
        ws_s.append([
            s["date"], s["day_n"], s["worktree"], s["started"], s["ended"],
            s["duration_min"], s["tool_calls"], s["tokens_in"], s["tokens_out"],
            s["tokens_cache_read"], s["tokens_saved"], s["cost_usd"],
            s["in_out_ratio"], s["branch"], s["commits"], s["source"],
        ])

    # 2. Tokens — Daily Rollup
    ws_d = reset_sheet("Tokens — Daily Rollup", [
        "Date", "Day-N", "Total Sessions", "Total Hours", "Total Tool Calls",
        "Total Tokens In", "Total Tokens Out", "Total Tokens Saved", "Total Cost USD",
    ])
    by_date = defaultdict(lambda: {"sessions": 0, "min": 0, "tc": 0, "in": 0, "out": 0, "saved": 0, "cost": 0.0})
    for s in sessions:
        if s["source"] == "<unknown>":
            continue
        d = by_date[s["date"]]
        d["sessions"] += 1
        d["min"] += s["duration_min"]
        d["tc"] += s["tool_calls"]
        d["in"] += s["tokens_in"]
        d["out"] += s["tokens_out"]
        d["saved"] += s["tokens_saved"]
        d["cost"] += s["cost_usd"]
    for date in sorted(by_date):
        d = by_date[date]
        ws_d.append([date, day_n(date), d["sessions"], round(d["min"] / 60, 2),
                     d["tc"], d["in"], d["out"], d["saved"], round(d["cost"], 4)])

    # 3. Tokens — Per-Chat Comparison
    ws_c = reset_sheet("Tokens — Per-Chat Comparison", [
        "Worktree", "Sessions", "Hours", "Tool Calls", "Tokens In",
        "Tokens Out", "Tokens Saved", "Cost USD", "Tokens Saved/Hour", "In:Out Ratio",
    ])
    by_wt = defaultdict(lambda: {"sessions": 0, "min": 0, "tc": 0, "in": 0, "out": 0, "saved": 0, "cost": 0.0})
    for s in sessions:
        if s["source"] == "<unknown>":
            continue
        w = by_wt[s["worktree"]]
        w["sessions"] += 1
        w["min"] += s["duration_min"]
        w["tc"] += s["tool_calls"]
        w["in"] += s["tokens_in"]
        w["out"] += s["tokens_out"]
        w["saved"] += s["tokens_saved"]
        w["cost"] += s["cost_usd"]
    grand = defaultdict(int)
    for wt in ("MAIN", "FE", "BE", "Cowork"):
        if wt not in by_wt:
            continue
        w = by_wt[wt]
        hrs = round(w["min"] / 60, 2)
        ratio = round(w["in"] / w["out"], 2) if w["out"] else 0
        saved_per_hr = round(w["saved"] / hrs) if hrs else 0
        ws_c.append([wt, w["sessions"], hrs, w["tc"], w["in"], w["out"],
                     w["saved"], round(w["cost"], 4), saved_per_hr, ratio])
        for k in ("sessions", "tc", "in", "out", "saved"):
            grand[k] += w[k]
        grand["min"] += w["min"]
        grand["cost"] += w["cost"]
    if grand:
        hrs = round(grand["min"] / 60, 2)
        ratio = round(grand["in"] / grand["out"], 2) if grand["out"] else 0
        saved_per_hr = round(grand["saved"] / hrs) if hrs else 0
        ws_c.append(["GRAND TOTAL", grand["sessions"], hrs, grand["tc"], grand["in"],
                     grand["out"], grand["saved"], round(grand["cost"], 4), saved_per_hr, ratio])

    # 4. Backfill_Notes
    ws_n = reset_sheet("Backfill_Notes", ["Date", "Worktree", "Field", "Reason", "Mitigation"])
    for n in notes:
        ws_n.append([n["date"], n["worktree"], n["field"], n["reason"], n["mitigation"]])

    wb.save(WORK_LOG)
    print(f"✓ Updated {WORK_LOG.relative_to(REPO_ROOT)}")
    print(f"  Sessions: {len(sessions)} ({sum(1 for s in sessions if s['source'] != '<unknown>')} with data)")
    print(f"  Backfill notes: {len(notes)}")


def main():
    sessions, notes = load_sessions()
    sessions, notes = add_fe_be_placeholder_rows(sessions, notes)
    write_workbook(sessions, notes)


if __name__ == "__main__":
    main()
